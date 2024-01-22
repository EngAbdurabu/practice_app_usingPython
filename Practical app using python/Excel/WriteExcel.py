import openpyxl
from pathlib import Path

# create Excel file
ExcelFile = openpyxl.Workbook()
print(ExcelFile.sheetnames)

# change sheet name
sheet = ExcelFile.active
sheet.title = "FirstSheet"

# create sheet
ExcelFile.create_sheet()
ExcelFile.create_sheet(title= "OneSheet")
# you can determine the name and index of sheet
ExcelFile.create_sheet(index=1, title="SecondSheet")

# delete sheet
del ExcelFile["OneSheet"]

# write to sheet
sheet = ExcelFile["SecondSheet"]
sheet["A1"] = "Hellow, World"
print(sheet["A1"].value)

# exercise
sheet2 = ExcelFile["FirstSheet"]
# sheet2["C1"] = "Ahmed"
# sheet2["C2"] = "Mohammed"
# sheet2["C3"] = "Saleh"

# another solution
names = ["Ahmed", "Ali", "Saleh", "Mohammed", "Majid"]
salery = ["1000$", "5000$", "6000$", "2000$", "1500$"]

for i in range(1, len(names) + 1):
	sheet2['A1'] = "names"
	sheet2["B1"] = "salery"
	sheet2.cell(row=i+1, column=1).value = names[i-1]
	sheet2.cell(row=i+1, column=2).value = salery[i-1]


# saving file
ExcelFile.save(filename=Path.home()/Path("OneDrive", "سطح المكتب",
                                          "python_practical")/"New_file.xlsx")
