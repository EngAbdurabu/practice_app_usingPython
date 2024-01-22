import openpyxl
from pathlib import Path

# open Excel file
excel_file = openpyxl.load_workbook(Path.home() / Path("OneDrive", "سطح المكتب",
                                                       "python_practical",
                                                       "file.xlsx"))

print(excel_file.sheetnames)  # to know the name of sheet in file
sheet1 = excel_file["Sheet1"]
print(sheet1.title)  # to display the title of sheet

activeSheet = excel_file.active
print(activeSheet.title)

print(sheet1["A1"].value)
print(sheet1["B1"].value)
print(sheet1["c1"].value)

print(sheet1['b2'].row)  # to know the row
print(sheet1['b2'].column)  # to know the columns

print(sheet1['C1'].coordinate)
print(sheet1.cell(row=3, column=2).value)
print("-" * 50)

for i in range(1, 7):
	print(i, ":", sheet1.cell(row=i, column=1).value)

print("-" * 50)
print(sheet1.max_row)
print(sheet1.max_column)
print("-" * 50)
total = 0
for i in range(1, 7):
	print(sheet1.cell(row=i, column=1).value,
	      sheet1.cell(row=i, column=2).value)
	total += sheet1.cell(row=i, column=2).value
#
print(f"The total salary of employees is : {total}")

print("-" * 50)
